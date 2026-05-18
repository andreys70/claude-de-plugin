#!/usr/bin/env python3
"""
Generate code annotation spreadsheet for a DM/DQ QuickETL conf file.

Two-phase pipeline:

  Phase 1 (skeleton):
     scan the conf, build the xlsx skeleton with header / identified-tables /
     line-numbered code rows, and emit an `annotation_context.json` listing
     every line that needs an LLM-generated note plus structural context
     (enclosing step, alias map, line-ranges for cross-reference).

  Phase 2 (merge):
     read an `annotations.json` produced by the LLM (list of {line, note})
     and write the notes into column D of the sheet.

Usage:
  # Phase 1 — emit skeleton + context
  python3 generate_annotation.py skeleton \
      --xlsx <path> --conf <path> --sheet <name> \
      --developer "<name>" --report-name "<conf basename>" \
      --overview "<2-4 sentence overview>" \
      --context-out <path>

  # Phase 2 — merge annotations into the sheet
  python3 generate_annotation.py merge \
      --xlsx <path> --sheet <name> --annotations <path>

  # Phase 0 (legacy quick-start) — skeleton + simple regex annotations
  python3 generate_annotation.py simple \
      --xlsx ... --conf ... --sheet ... --developer ... \
      --report-name ... --overview ...
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Install with:", file=sys.stderr)
    print("  pip3 install openpyxl --break-system-packages --quiet", file=sys.stderr)
    sys.exit(1)


# ============================================================================
# Identified-tables extraction
# ============================================================================

ALLOWED_SCHEMA_RE = re.compile(
    r'^('
    r'finance_mm_(sandbox|dm|sox)|'
    r'finance_(sandbox|dm|sox)|'
    r'finance_sandbox_stg|'
    r'ued_[\w]+_(prd_dwh|sox_dwh|dwh|stg)|'
    r'risk_analytics_[\w]+|'
    r'finance_[\w]+_dm|'
    r'finance_[\w]+_sox'
    r')$'
)


def extract_identified_tables(conf_text: str) -> list[str]:
    tables = set()
    pat = re.compile(r'\b([a-z][\w]*)\.([A-Za-z][\w]*)\b')
    for m in pat.finditer(conf_text):
        schema, table = m.group(1), m.group(2)
        if not ALLOWED_SCHEMA_RE.match(schema):
            continue
        if len(table) < 3:
            continue
        tables.add(f"{schema}.{table}")
    for token in ("RPT_SOX_SETUP", "RPT_SOX_METADATA", "RPT_SOX_COMPLETENESS",
                  "RPT_SOX_ACCURACY", "RPT_SOX_AUDIT_RESULTS"):
        if re.search(rf'\b{token}\b', conf_text):
            tables.add(token)
    return sorted(tables)


# ============================================================================
# Structural parsing of HOCON / SQL conf
# ============================================================================

def parse_structure(lines: list[str]) -> dict:
    """Walk the conf line-by-line and build a structural index.

    Returns:
        {
          'steps': {step_name: (start_line, end_line)},   # 1-indexed inclusive
          'sql_blocks': [{step, start, end}],             # SQL between triple-quotes
          'aliases': {alias: source},                     # alias resolution map
        }
    """
    steps: dict = {}
    sql_blocks: list = []
    aliases: dict = {}

    current_step = None
    step_start = None
    step_brace_depth = 0
    in_step_body = False

    in_sql_block = False
    sql_block_start = None
    sql_block_step = None

    step_re = re.compile(r'^\s*([A-Za-z_][\w]*)\s*=\s*\{\s*$')
    triple_quote_re = re.compile(r'"""')

    for i, line in enumerate(lines, start=1):
        if triple_quote_re.search(line):
            if not in_sql_block:
                in_sql_block = True
                sql_block_start = i
                sql_block_step = current_step
            else:
                sql_blocks.append({
                    'step': sql_block_step,
                    'start': sql_block_start,
                    'end': i,
                })
                in_sql_block = False
                sql_block_start = None
            continue

        if not in_sql_block:
            m = step_re.match(line)
            if m and current_step is None:
                current_step = m.group(1)
                step_start = i
                step_brace_depth = 1
                in_step_body = True
                continue
            if in_step_body:
                step_brace_depth += line.count('{') - line.count('}')
                if step_brace_depth <= 0:
                    steps[current_step] = (step_start, i)
                    current_step = None
                    in_step_body = False

        if in_sql_block:
            for jm in re.finditer(
                r'\b(FROM|JOIN|INNER\s+JOIN|LEFT\s+JOIN|RIGHT\s+JOIN|FULL\s+OUTER\s+JOIN|CROSS\s+JOIN)\s+'
                r'([\w.${}"]+)\s+([a-zA-Z_][\w]*)\b',
                line, re.IGNORECASE
            ):
                source = jm.group(2)
                alias = jm.group(3)
                if alias.upper() in ('ON', 'WHERE', 'AS', 'AND', 'OR', 'LIMIT'):
                    continue
                aliases[alias] = source

    return {'steps': steps, 'sql_blocks': sql_blocks, 'aliases': aliases}


# ============================================================================
# Annotation context builder
# ============================================================================

def line_kind(line: str) -> str:
    s = line.strip()
    if not s:
        return 'blank'
    if s.startswith('//') or s.startswith('#'):
        return 'comment'
    if re.match(r'^\s*include\s+required', line):
        return 'include'
    if re.search(r'"""', s):
        return 'sql_boundary'
    if s in ('}', '},', ']', '],', ')'):
        return 'block_close'
    if re.match(r'^\s*[A-Za-z_][\w-]*\s*=\s*\{|^\s*\{', line) and '"' not in line:
        return 'section_open'
    if re.match(r'^\s*[a-zA-Z_][\w-]*\s*=', line):
        return 'kv'
    su = s.upper()
    if su.startswith(('SELECT', 'FROM', 'WHERE', 'GROUP BY', 'ORDER BY',
                      'UNION', 'WITH', 'INNER JOIN', 'LEFT JOIN', 'RIGHT JOIN',
                      'FULL OUTER JOIN', 'JOIN', 'CROSS JOIN', 'INSERT INTO',
                      'INSERT OVERWRITE', 'CASE', 'WHEN', 'THEN', 'ELSE', 'END',
                      'HAVING', 'IF', 'EXISTS', 'AND', 'OR', 'NOT')):
        return 'sql'
    return 'other'


def build_annotation_context(lines: list[str], structure: dict) -> list[dict]:
    contexts = []

    def find_step(i):
        for name, (s, e) in structure['steps'].items():
            if s <= i <= e:
                return name
        return None

    for i, line in enumerate(lines, start=1):
        ctx = {
            'line': i,
            'code': line,
            'kind': line_kind(line),
            'step': find_step(i),
        }
        contexts.append(ctx)
    return contexts


# ============================================================================
# Skeleton xlsx writer
# ============================================================================

def _styles():
    return {
        'title_font': Font(bold=True, size=14, color="FFFFFF"),
        'title_fill': PatternFill("solid", fgColor="305496"),
        'header_font': Font(bold=True, size=11),
        'table_header_font': Font(bold=True, size=11, color="FFFFFF"),
        'table_header_fill': PatternFill("solid", fgColor="4472C4"),
        'code_font': Font(name="Courier New", size=10),
        'note_font': Font(name="Calibri", size=10),
        'wrap': Alignment(wrap_text=True, vertical="top"),
    }


def write_skeleton(wb, sheet_name, conf_path, developer, report_name, overview):
    with open(conf_path, 'r') as f:
        conf_text = f.read()
    lines = conf_text.split('\n')

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    s = _styles()

    c = ws.cell(row=1, column=2, value="Code Reviewer")
    c.font = s['title_font']; c.fill = s['title_fill']
    c = ws.cell(row=1, column=3, value="Purpose: Document understanding of code in meeting the business purpose")
    c.font = s['title_font']; c.fill = s['title_fill']
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=10)

    ws.cell(row=5, column=2, value="Developer:").font = s['header_font']
    ws.cell(row=5, column=3, value=developer)
    ws.cell(row=6, column=2, value="Date:").font = s['header_font']
    ws.cell(row=6, column=3, value=datetime.now().strftime("%Y-%m-%d"))
    ws.cell(row=7, column=2, value="Report Name:").font = s['header_font']
    ws.cell(row=7, column=3, value=report_name)
    ws.cell(row=8, column=2, value="Brief Overview of Code").font = s['header_font']
    c = ws.cell(row=8, column=3, value=overview)
    c.alignment = s['wrap']
    ws.merge_cells(start_row=8, start_column=3, end_row=8, end_column=10)

    tables = extract_identified_tables(conf_text)
    ws.cell(row=10, column=2, value="Query").font = s['table_header_font']
    ws.cell(row=10, column=2).fill = s['table_header_fill']
    ws.cell(row=10, column=3, value="Identified tables").font = s['table_header_font']
    ws.cell(row=10, column=3).fill = s['table_header_fill']
    ws.cell(row=10, column=4, value="Tech Team Notes").font = s['table_header_font']
    ws.cell(row=10, column=4).fill = s['table_header_fill']
    row = 11
    for t in tables:
        ws.cell(row=row, column=2, value=1.0)
        ws.cell(row=row, column=3, value=t).font = s['code_font']
        row += 1

    row += 2

    ws.cell(row=row, column=2, value="Line Number").font = s['table_header_font']
    ws.cell(row=row, column=2).fill = s['table_header_fill']
    ws.cell(row=row, column=3, value="Original code query").font = s['table_header_font']
    ws.cell(row=row, column=3).fill = s['table_header_fill']
    ws.cell(row=row, column=4, value="Developer Notes").font = s['table_header_font']
    ws.cell(row=row, column=4).fill = s['table_header_fill']
    ws.cell(row=row, column=5, value="PO Notes").font = s['table_header_font']
    ws.cell(row=row, column=5).fill = s['table_header_fill']
    row += 1

    ws.cell(row=row, column=2, value="Query 1")
    ws.cell(row=row, column=4, value="--")
    row += 1

    for i, line in enumerate(lines, start=1):
        ws.cell(row=row, column=2, value=i)
        c = ws.cell(row=row, column=3, value=line)
        c.font = s['code_font']
        c.alignment = Alignment(vertical='top')
        row += 1

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 100
    ws.column_dimensions['D'].width = 70
    ws.column_dimensions['E'].width = 30


def merge_annotations(wb, sheet_name, annotations: list[dict]):
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    s = _styles()

    line_to_row = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, int):
            line_to_row[v] = r

    by_line = {a['line']: a['note'] for a in annotations}

    written = 0
    for line_num, xlsx_row in line_to_row.items():
        note = by_line.get(line_num, '--')
        c = ws.cell(row=xlsx_row, column=4, value=note)
        c.font = s['note_font']
        c.alignment = s['wrap']
        written += 1
    return written


# ============================================================================
# Simple regex-based annotator (fallback / quick mode)
# ============================================================================

def simple_annotate(line: str) -> str:
    s = line.strip()
    if not s:
        return '--'
    if s.startswith('//') or s.startswith('#'):
        return 'This is a comment line'
    m = re.match(r'^\s*include\s+required\s*\(\s*file\s*\(\s*"([^"]+)"\s*\)\s*\)', line)
    if m:
        return f'Include the file "{m.group(1)}"'
    m = re.match(r'^\s*([A-Za-z_][\w-]*)\s*=\s*\{', line)
    if m:
        return f'Set the following attribute values for the configuration {m.group(1)}:'
    if s in ('}', '},', ']', '],', ')'):
        return '--'
    if re.match(r'^\s*steps\s*=\s*\[', line):
        return '- steps, is set to the values mentioned in following lines, and defined in the code following:'
    m = re.match(r'^\s*([A-Za-z_][\w-]*)\s*=\s*"([^"]*)"\s*,?\s*$', line)
    if m:
        return f'- {m.group(1)} is set to "{m.group(2)}"'
    m = re.match(r'^\s*([A-Za-z_][\w-]*)\s*=\s*([^\s,{]+)\s*,?\s*$', line)
    if m:
        return f'- {m.group(1)} is set to {m.group(2)}'
    if '"""' in s:
        return 'SQL block boundary (triple-quote)'
    return f'Code line: {s[:200]}'


# ============================================================================
# CLI
# ============================================================================

def cmd_skeleton(args):
    try:
        wb = openpyxl.load_workbook(args.xlsx)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    if len(args.sheet) > 31:
        print(f"ERROR: sheet name '{args.sheet}' exceeds Excel's 31-char limit. "
              f"Abbreviate (e.g. _loanpro_->_lp_, _transaction->_txn).", file=sys.stderr)
        sys.exit(1)

    write_skeleton(wb, args.sheet, args.conf, args.developer,
                   args.report_name, args.overview)
    wb.save(args.xlsx)

    with open(args.conf, 'r') as f:
        conf_text = f.read()
    lines = conf_text.split('\n')
    structure = parse_structure(lines)
    contexts = build_annotation_context(lines, structure)

    payload = {
        'conf_path': str(Path(args.conf).resolve()),
        'sheet_name': args.sheet,
        'xlsx_path': str(Path(args.xlsx).resolve()),
        'total_lines': len(lines),
        'structure': {
            'steps': {k: list(v) for k, v in structure['steps'].items()},
            'sql_blocks': structure['sql_blocks'],
            'aliases': structure['aliases'],
        },
        'lines': contexts,
    }
    with open(args.context_out, 'w') as f:
        json.dump(payload, f, indent=2)

    print(f"OK: skeleton written to {args.xlsx} [sheet: {args.sheet}]")
    print(f"OK: annotation context written to {args.context_out} ({len(lines)} lines)")
    print(f"\nNext: have the LLM agent read {args.context_out} and emit annotations.json")
    print(f"Then run: {sys.argv[0]} merge --xlsx {args.xlsx} --sheet {args.sheet} --annotations <path>")


def cmd_merge(args):
    wb = openpyxl.load_workbook(args.xlsx)
    with open(args.annotations, 'r') as f:
        data = json.load(f)
    annotations = data['annotations'] if isinstance(data, dict) else data
    written = merge_annotations(wb, args.sheet, annotations)
    wb.save(args.xlsx)
    print(f"OK: merged {len(annotations)} annotations into {args.xlsx} [sheet: {args.sheet}]")
    print(f"OK: {written} rows updated")


def cmd_simple(args):
    try:
        wb = openpyxl.load_workbook(args.xlsx)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    if len(args.sheet) > 31:
        print(f"ERROR: sheet name '{args.sheet}' exceeds 31-char limit", file=sys.stderr)
        sys.exit(1)

    write_skeleton(wb, args.sheet, args.conf, args.developer,
                   args.report_name, args.overview)

    with open(args.conf, 'r') as f:
        lines = f.read().split('\n')
    annotations = [{'line': i, 'note': simple_annotate(L)}
                   for i, L in enumerate(lines, start=1)]
    merge_annotations(wb, args.sheet, annotations)
    wb.save(args.xlsx)
    print(f"OK: simple-mode annotation written to {args.xlsx} [sheet: {args.sheet}]")


def main():
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest='cmd', required=True)

    sp_skel = sub.add_parser('skeleton', help='Phase 1: skeleton + context')
    sp_skel.add_argument('--xlsx', required=True)
    sp_skel.add_argument('--conf', required=True)
    sp_skel.add_argument('--sheet', required=True)
    sp_skel.add_argument('--developer', default='')
    sp_skel.add_argument('--report-name', required=True)
    sp_skel.add_argument('--overview', required=True)
    sp_skel.add_argument('--context-out', required=True)
    sp_skel.set_defaults(func=cmd_skeleton)

    sp_merge = sub.add_parser('merge', help='Phase 2: merge annotations.json')
    sp_merge.add_argument('--xlsx', required=True)
    sp_merge.add_argument('--sheet', required=True)
    sp_merge.add_argument('--annotations', required=True)
    sp_merge.set_defaults(func=cmd_merge)

    sp_simple = sub.add_parser('simple', help='Quick mode: regex-based notes')
    sp_simple.add_argument('--xlsx', required=True)
    sp_simple.add_argument('--conf', required=True)
    sp_simple.add_argument('--sheet', required=True)
    sp_simple.add_argument('--developer', default='')
    sp_simple.add_argument('--report-name', required=True)
    sp_simple.add_argument('--overview', required=True)
    sp_simple.set_defaults(func=cmd_simple)

    args = parser.parse_args()
    args.func(args)


if __name__ == '__main__':
    main()
